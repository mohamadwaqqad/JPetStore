from openpyxl import load_workbook


class Utils:
    @staticmethod
    def read_data_from_excal(sheet):
        data_list = []
        wb = load_workbook(filename=r"./testdata/testdata.xlsx")
        sh = wb[sheet]

        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list
